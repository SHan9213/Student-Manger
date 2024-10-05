import openpyxl

class Result:
    def __init__(self, student):
        self.student = student

    def calculate_average(self):
        marks = self.student.get_marks().values()
        return sum(marks) / len(marks)

    def determine_result(self):
        average = self.calculate_average()
        return "Fail" if average < 33 else "Pass"

    def save_to_excel(self, file_name='student_results.xlsx'):
        try:
            workbook = openpyxl.load_workbook(file_name)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        sheet = workbook.active
        if sheet.max_row == 1:
            sheet.append(["Student Name", "Physics", "Chemistry", "Maths", "Computer Science", "Average", "Result"])
        sheet.append([
            self.student.name,
            self.student.physics,
            self.student.chemistry,
            self.student.maths,
            self.student.computer_science,
            self.calculate_average(),
            self.determine_result()
        ])

        # Save the workbook
        workbook.save(file_name)
        print(f"Data saved to {file_name}")
